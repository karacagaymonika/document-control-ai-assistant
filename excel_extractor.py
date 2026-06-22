from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook


LABEL_ALIASES = {
    "document_number": {
        "document number",
        "document no",
        "document no.",
        "doc number",
        "doc no",
        "document reference",
        "document ref",
        "reference number",
        "reference no",
    },
    "title": {"title", "document title", "document name"},
    "revision": {"revision", "rev", "current revision"},
    "status": {
        "status",
        "document status",
        "issue status",
        "suitability",
        "suitability code",
    },
    "project": {"project", "project name"},
    "discipline": {"discipline", "department"},
    "document_type": {"document type", "type", "document category"},
    "originator": {
        "originator",
        "originating company",
        "organisation",
        "organization",
        "company",
    },
    "issue_date": {
        "issue date",
        "date issued",
        "issued date",
        "created date",
        "document date",
    },
    "prepared_by": {"prepared by", "author", "created by"},
    "checked_by": {"checked by", "reviewed by"},
    "approved_by": {"approved by", "authorised by", "authorized by"},
    "owner": {"owner", "document owner"},
    "due_date": {"due date", "response due date"},
}

ALIAS_TO_FIELD = {
    alias: field
    for field, aliases in LABEL_ALIASES.items()
    for alias in aliases
}


def _normalise_label(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[\s_\-/]+", " ", text)
    return re.sub(r"[.:]+$", "", text).strip()


def _clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_known_label(value: Any) -> bool:
    return _normalise_label(value) in ALIAS_TO_FIELD


def _title_candidate(worksheet) -> str:
    candidates: list[tuple[float, str]] = []
    max_row = min(worksheet.max_row, 15)
    max_column = min(worksheet.max_column, 15)

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_column,
    ):
        for cell in row:
            value = _clean_value(cell.value)
            if not value or _is_known_label(value):
                continue
            if len(value) < 8 or len(value) > 180:
                continue

            font_size = float(cell.font.sz or 10)
            bold_bonus = 5 if cell.font.bold else 0
            letter_count = sum(character.isalpha() for character in value)
            uppercase_ratio = (
                sum(character.isupper() for character in value if character.isalpha())
                / max(1, letter_count)
            )

            merged_bonus = 0
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    merged_bonus = 8 + min(
                        8,
                        (merged_range.max_col - merged_range.min_col)
                        + (merged_range.max_row - merged_range.min_row),
                    )
                    break

            score = (
                font_size
                + bold_bonus
                + merged_bonus
                + uppercase_ratio * 4
                + min(len(value) / 20, 6)
            )
            candidates.append((score, value))

    candidates.sort(reverse=True)
    return candidates[0][1] if candidates else ""


def _standardise_date(value: str) -> str:
    if not value:
        return ""
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return value
    return parsed.date().isoformat()


def _first_non_empty_sheet(workbook):
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        non_empty = 0
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row, 30),
            min_col=1,
            max_col=min(worksheet.max_column, 15),
        ):
            non_empty += sum(1 for cell in row if _clean_value(cell.value))
        if non_empty:
            return worksheet
    return workbook[workbook.sheetnames[0]]


def extract_excel_metadata(file_bytes: bytes) -> dict[str, Any]:
    """Extract controlled-document metadata from one Excel workbook.

    The function is intentionally rule-based and transparent. Every extracted
    value must still be reviewed by the user before it is saved.
    """
    workbook = load_workbook(
        BytesIO(file_bytes),
        data_only=True,
        read_only=False,
    )
    worksheet = _first_non_empty_sheet(workbook)

    metadata: dict[str, str] = {}
    evidence: dict[str, dict[str, str]] = {}

    max_row = min(worksheet.max_row, 80)
    max_column = min(worksheet.max_column, 24)

    for row_number in range(1, max_row + 1):
        for column_number in range(1, max_column + 1):
            label_cell = worksheet.cell(row_number, column_number)
            field = ALIAS_TO_FIELD.get(_normalise_label(label_cell.value))
            if not field:
                continue

            extracted_value = ""
            value_coordinate = ""

            for offset in range(1, 5):
                target_column = column_number + offset
                if target_column > max_column:
                    break
                value_cell = worksheet.cell(row_number, target_column)
                candidate = _clean_value(value_cell.value)
                if candidate and not _is_known_label(candidate):
                    extracted_value = candidate
                    value_coordinate = value_cell.coordinate
                    break

            if not extracted_value:
                for offset in range(1, 3):
                    target_row = row_number + offset
                    if target_row > max_row:
                        break
                    value_cell = worksheet.cell(target_row, column_number)
                    candidate = _clean_value(value_cell.value)
                    if candidate and not _is_known_label(candidate):
                        extracted_value = candidate
                        value_coordinate = value_cell.coordinate
                        break

            if extracted_value and not metadata.get(field):
                metadata[field] = extracted_value
                evidence[field] = {
                    "label_cell": label_cell.coordinate,
                    "value_cell": value_coordinate,
                    "sheet": worksheet.title,
                }

    visual_title = _title_candidate(worksheet)
    title_evidence = evidence.get("title", {})
    label_coordinate = title_evidence.get("label_cell", "")
    label_row_match = re.search(r"\d+", label_coordinate)
    labelled_title_row = int(label_row_match.group()) if label_row_match else 999

    if visual_title and (
        not metadata.get("title")
        or labelled_title_row > 15
    ):
        metadata["title"] = visual_title
        evidence["title"] = {
            "label_cell": "",
            "value_cell": "document header",
            "sheet": worksheet.title,
        }

    metadata["issue_date"] = _standardise_date(metadata.get("issue_date", ""))
    metadata["due_date"] = _standardise_date(metadata.get("due_date", ""))

    if metadata.get("status"):
        metadata["status"] = re.sub(r"\s+", " ", metadata["status"]).strip()

    core_fields = [
        "document_number",
        "title",
        "revision",
        "status",
        "project",
        "discipline",
    ]
    supporting_fields = [
        "document_type",
        "originator",
        "issue_date",
        "prepared_by",
        "checked_by",
        "approved_by",
    ]

    confidence = min(
        100,
        sum(bool(metadata.get(field)) for field in core_fields) * 12
        + sum(bool(metadata.get(field)) for field in supporting_fields) * 4,
    )

    warnings: list[str] = []
    missing_core = [
        field
        for field in core_fields
        if not metadata.get(field)
    ]
    if missing_core:
        warnings.append(
            "Please review missing key fields: "
            + ", ".join(
                field.replace("_", " ").title()
                for field in missing_core
            )
            + "."
        )

    if len(workbook.sheetnames) > 1:
        warnings.append(
            f"The workbook contains {len(workbook.sheetnames)} sheets. "
            f"Metadata was extracted from '{worksheet.title}'."
        )

    preview_rows: list[list[str]] = []
    preview_max_row = min(worksheet.max_row, 25)
    preview_max_column = min(worksheet.max_column, 10)

    for row_number in range(1, preview_max_row + 1):
        row_values = [
            _clean_value(worksheet.cell(row_number, column_number).value)
            for column_number in range(1, preview_max_column + 1)
        ]
        if any(row_values):
            preview_rows.append(row_values)

    preview = pd.DataFrame(preview_rows)

    return {
        "metadata": metadata,
        "evidence": evidence,
        "confidence": int(confidence),
        "warnings": warnings,
        "sheet_name": worksheet.title,
        "sheet_names": list(workbook.sheetnames),
        "preview": preview,
    }
